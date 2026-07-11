from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.excel_crypto import (
  ExcelPasswordError,
  is_encrypted_excel,
  load_workbook_from_path,
  save_workbook_to_path,
)
from app.services.excel_service import append_csv_to_excel
from app.services.header_import import parse_headers_from_text


class ExcelPasswordTests(unittest.TestCase):
  def test_parse_tab_separated(self):
    headers = parse_headers_from_text('Col A\tCol B\tCol C')
    self.assertEqual(headers, ['Col A', 'Col B', 'Col C'])

  def test_edit_protection_opens_without_password(self):
    with tempfile.TemporaryDirectory() as tmp:
      path = Path(tmp) / 'editavel.xlsx'
      workbook = Workbook()
      sheet = workbook.active
      sheet.title = 'Julho 2026'
      sheet.cell(row=1, column=1, value='A')
      save_workbook_to_path(workbook, path, password='edit123')

      self.assertFalse(is_encrypted_excel(path))
      loaded = load_workbook_from_path(path)
      self.assertTrue(loaded['Julho 2026'].protection.sheet)

  def test_append_to_new_file_with_edit_password(self):
    with tempfile.TemporaryDirectory() as tmp:
      folder = Path(tmp)
      excel_path = folder / 'saida.xlsx'
      csv_path = folder / 'entrada.csv'
      csv_path.write_text('Item\t10\n', encoding='utf-8')

      sheet_name, count = append_csv_to_excel(
        csv_path,
        excel_path,
        ['Nome', 'Qtd'],
        excel_password='edit123',
      )
      self.assertEqual(count, 1)
      self.assertEqual(sheet_name, 'Julho 2026')
      self.assertFalse(is_encrypted_excel(excel_path))

      loaded = load_workbook(excel_path)
      self.assertTrue(loaded['Julho 2026'].protection.sheet)
      self.assertEqual(loaded['Julho 2026'].cell(row=2, column=1).value, 'Item')

  def test_append_again_without_open_password(self):
    with tempfile.TemporaryDirectory() as tmp:
      folder = Path(tmp)
      excel_path = folder / 'saida.xlsx'
      csv_path = folder / 'entrada.csv'
      csv_path.write_text('A\t1\n', encoding='utf-8')
      append_csv_to_excel(csv_path, excel_path, ['Nome', 'Qtd'], excel_password='edit123')

      csv_path.write_text('B\t2\n', encoding='utf-8')
      append_csv_to_excel(csv_path, excel_path, ['Nome', 'Qtd'], excel_password='edit123')

      loaded = load_workbook(excel_path)
      self.assertEqual(loaded['Julho 2026'].cell(row=3, column=1).value, 'B')

  def test_legacy_encrypted_file_still_needs_open_password(self):
    with tempfile.TemporaryDirectory() as tmp:
      path = Path(tmp) / 'legado.xlsx'
      workbook = Workbook()
      workbook.active['A1'] = 'x'

      import io
      from msoffcrypto.format.ooxml import OOXMLFile

      plain = io.BytesIO()
      workbook.save(plain)
      plain.seek(0)
      encrypted = io.BytesIO()
      OOXMLFile(plain).encrypt('abrir123', encrypted)
      path.write_bytes(encrypted.getvalue())

      self.assertTrue(is_encrypted_excel(path))
      with self.assertRaises(ExcelPasswordError):
        load_workbook_from_path(path)
      loaded = load_workbook_from_path(path, password='abrir123')
      self.assertEqual(loaded.active['A1'].value, 'x')


if __name__ == '__main__':
  unittest.main()
