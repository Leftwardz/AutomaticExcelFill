from __future__ import annotations

import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app.models.schema import AppConfig, Flow
from app.services.csv_batch import collect_ready_csv_batch, is_active_watch_file
from app.services.excel_service import append_csv_to_excel


class CsvBatchTests(unittest.TestCase):
  JULY_2026 = datetime(2026, 7, 10)

  def test_collect_ready_csv_batch_groups_matching_files(self):
    flow = Flow(
      name='Lote',
      source_filename='dados_*.csv',
      excel_directory='/tmp',
      excel_filename='out.xlsx',
    )
    config = AppConfig(
      watch_folder='/tmp/watch',
      processed_subfolder='processados',
      failed_subfolder='falhas',
    )
    with tempfile.TemporaryDirectory() as tmp:
      watch_folder = Path(tmp)
      config.watch_folder = str(watch_folder)
      (watch_folder / 'dados_a.csv').write_text('A\t1\n', encoding='utf-8')
      (watch_folder / 'dados_b.csv').write_text('B\t2\n', encoding='utf-8')
      (watch_folder / 'outros.csv').write_text('X\t9\n', encoding='utf-8')
      (watch_folder / 'processados').mkdir()
      (watch_folder / 'processados' / 'dados_old.csv').write_text('Z\t0\n', encoding='utf-8')

      batch = collect_ready_csv_batch(
        watch_folder,
        flow,
        config,
        skip_stability_wait=True,
      )
      self.assertEqual([path.name for path in batch], ['dados_a.csv', 'dados_b.csv'])

  def test_is_active_watch_file_ignores_processed_subfolder(self):
    config = AppConfig(processed_subfolder='processados', failed_subfolder='falhas')
    with tempfile.TemporaryDirectory() as tmp:
      watch_folder = Path(tmp)
      active = watch_folder / 'dados.csv'
      active.write_text('A\n', encoding='utf-8')
      processed_dir = watch_folder / 'processados'
      processed_dir.mkdir()
      processed = processed_dir / 'dados.csv'
      processed.write_text('A\n', encoding='utf-8')
      self.assertTrue(is_active_watch_file(active, watch_folder, config))
      self.assertFalse(is_active_watch_file(processed, watch_folder, config))

  def test_append_csv_to_excel_accepts_multiple_files(self):
    with tempfile.TemporaryDirectory() as tmp:
      first = Path(tmp) / 'a.csv'
      second = Path(tmp) / 'b.csv'
      excel_path = Path(tmp) / 'saida.xlsx'
      first.write_text('A\t1\n', encoding='utf-8')
      second.write_text('B\t2\n', encoding='utf-8')
      append_csv_to_excel(
        first,
        excel_path,
        ['Col1', 'Col2'],
        extra_csv_paths=[second],
        sheet_name='Julho',
        processed_on=self.JULY_2026,
      )
      workbook = load_workbook(excel_path)
      sheet = workbook['Julho']
      self.assertEqual(sheet.cell(row=2, column=1).value, 'A')
      self.assertEqual(sheet.cell(row=3, column=1).value, 'B')


if __name__ == '__main__':
  unittest.main()
