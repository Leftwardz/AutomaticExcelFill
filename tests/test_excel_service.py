from __future__ import annotations

import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, Side

from app.models.schema import Flow
from app.services.excel_service import (
  CELL_BORDER,
  GENERAL_NUMBER_FORMAT,
  DATE_NUMBER_FORMAT,
  CENTER_ALIGNMENT,
  DUPLICATE_FILL,
  HEADER_FILL,
  HEADER_FONT,
  META_SHEET_NAME,
  ROW_FILL_PURPLE,
  ROW_FILL_WHITE,
  TEXT_NUMBER_FORMAT,
  append_csv_to_excel,
  color_period_anchor,
  current_month_sheet_name,
  excel_full_path,
  read_tab_csv,
)


class ExcelServiceTests(unittest.TestCase):
  JULY_2026 = datetime(2026, 7, 10)
  SHEET_JULY = 'Julho'

  def test_current_month_sheet_name(self):
    self.assertEqual(current_month_sheet_name(self.JULY_2026), 'Julho')

  def test_excel_full_path_uses_year_subfolder(self):
    path = excel_full_path('/dados/excel', 'relatorio', when=self.JULY_2026)
    self.assertEqual(path, Path('/dados/excel/2026/relatorio.xlsx'))

  def test_append_creates_sheet_and_headers(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('A\tB\tC\n1\t2\t3\n4\t5\t6\n', encoding='utf-8')
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2', 'Col3']
      sheet_name, count = append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )
      self.assertEqual(sheet_name, self.SHEET_JULY)
      self.assertEqual(count, 3)
      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertEqual(sheet.cell(row=1, column=1).value, 'Col1')
      self.assertEqual(sheet.cell(row=2, column=1).value, 'A')
      self.assertEqual(sheet.cell(row=4, column=3).value, '6')

  def test_text_columns_ignore_number_stored_as_text_warning(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('12345\ttexto\n', encoding='utf-8')
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Codigo', 'Nome']
      append_csv_to_excel(
        csv_path, excel_path, headers, column_types=['text', 'text'],
        sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      import zipfile
      from xml.etree import ElementTree as ET
      with zipfile.ZipFile(excel_path, 'r') as zin:
        xml = zin.read('xl/worksheets/sheet1.xml')
      root = ET.fromstring(xml)
      ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
      ignored = next(
        element for element in root if element.tag == f'{{{ns}}}ignoredErrors'
      )
      error = next(
        child for child in ignored if child.tag == f'{{{ns}}}ignoredError'
      )
      self.assertEqual(error.get('numberStoredAsText'), '1')
      self.assertIn('A1:A1048576', error.get('sqref', ''))

  def test_duplicate_column_values_are_highlighted(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Codigo', 'Nome']
      csv_path.write_text('XYZ\tUnico\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, column_types=['text', 'text'],
        column_duplicate_checks=[True, False],
        sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      csv_path.write_text('ABC\tPrimeiro\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, column_types=['text', 'text'],
        column_duplicate_checks=[True, False],
        sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )
      csv_path.write_text('ABC\tSegundo\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, column_types=['text', 'text'],
        column_duplicate_checks=[True, False],
        sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertEqual(sheet.cell(row=3, column=1).fill.fgColor.rgb, DUPLICATE_FILL.fgColor.rgb)
      self.assertEqual(sheet.cell(row=4, column=1).fill.fgColor.rgb, DUPLICATE_FILL.fgColor.rgb)
      self.assertNotEqual(sheet.cell(row=2, column=2).fill.fgColor.rgb, DUPLICATE_FILL.fgColor.rgb)

  def test_standard_data_and_header_fonts(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('ABC\t123\n', encoding='utf-8')
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Codigo', 'Valor']
      append_csv_to_excel(
        csv_path, excel_path, headers, column_types=['text', 'number'],
        sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertEqual(sheet.cell(row=1, column=1).font.name, 'Calibri')
      self.assertEqual(sheet.cell(row=1, column=1).font.size, 11)
      self.assertTrue(sheet.cell(row=1, column=1).font.bold)
      self.assertEqual(sheet.cell(row=2, column=1).font.name, 'Calibri')
      self.assertEqual(sheet.cell(row=2, column=1).font.size, 8)
      self.assertEqual(sheet.cell(row=2, column=2).font.size, 8)

  def test_existing_sheet_gets_standard_font_on_new_rows(self):
    with tempfile.TemporaryDirectory() as tmp:
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']
      workbook = Workbook()
      sheet = workbook.active
      sheet.title = 'Julho'
      sheet['A1'] = 'Col1'
      sheet['B1'] = 'Col2'
      sheet['A2'] = 'Existente'
      sheet['B2'] = '123'
      sheet['A2'].font = Font(name='Arial', size=10)
      sheet['B2'].font = Font(name='Arial', size=10)
      workbook.save(excel_path)
      workbook.close()

      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('Novo\t456\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, column_types=['text', 'number'],
        sheet_name='Julho', processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook['Julho']
      self.assertEqual(sheet.cell(row=2, column=1).font.name, 'Calibri')
      self.assertEqual(sheet.cell(row=2, column=1).font.size, 8)
      self.assertEqual(sheet.cell(row=3, column=1).font.name, 'Calibri')
      self.assertEqual(sheet.cell(row=3, column=1).font.size, 8)

  def test_existing_sheet_layout_migrated_once(self):
    ugly_border = Border(
      left=Side(style='thick', color='FF0000'),
      right=Side(style='thick', color='FF0000'),
      top=Side(style='thick', color='FF0000'),
      bottom=Side(style='thick', color='FF0000'),
    )
    with tempfile.TemporaryDirectory() as tmp:
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']
      workbook = Workbook()
      sheet = workbook.active
      sheet.title = 'Julho'
      sheet['A1'] = 'Col1'
      sheet['B1'] = 'Col2'
      sheet['A2'] = 'Antigo'
      sheet['B2'] = 'Linha'
      for row_index in (1, 2):
        for col_index in (1, 2):
          sheet.cell(row=row_index, column=col_index).border = ugly_border
      workbook.save(excel_path)
      workbook.close()

      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('Novo\tValor\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name='Julho', processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook['Julho']
      self.assertFalse(sheet.sheet_view.showGridLines)
      self.assertEqual(sheet.cell(row=2, column=1).border.left.color.rgb, CELL_BORDER.left.color.rgb)
      self.assertNotEqual(sheet.cell(row=2, column=1).border.left.style, 'thick')
      meta = workbook[META_SHEET_NAME]
      self.assertEqual(meta.cell(row=2, column=4).value, 1)

      csv_path.write_text('Outro\tItem\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name='Julho', processed_on=self.JULY_2026,
      )
      workbook = load_workbook(excel_path)
      self.assertEqual(workbook[META_SHEET_NAME].cell(row=2, column=4).value, 1)
      self.assertEqual(workbook['Julho'].cell(row=3, column=1).value, 'Novo')
      self.assertEqual(workbook['Julho'].cell(row=4, column=1).value, 'Outro')

  def test_append_to_existing_sheet_with_legacy_lowercase_tab(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('A\tB\n', encoding='utf-8')
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']

      from openpyxl import Workbook
      workbook = Workbook()
      workbook.active.title = 'julho'
      workbook.save(excel_path)

      sheet_name, count = append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name='Julho', processed_on=self.JULY_2026,
      )
      self.assertEqual(sheet_name, 'julho')
      self.assertEqual(count, 1)

      workbook = load_workbook(excel_path)
      self.assertIn('julho', workbook.sheetnames)
      self.assertEqual(workbook['julho'].cell(row=2, column=1).value, 'A')

  def test_append_to_existing_sheet(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('X\tY\n', encoding='utf-8')
      excel_path = Path(tmp) / 'saida.xlsx'
      append_csv_to_excel(
        csv_path, excel_path, ['Col1', 'Col2'], sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )
      csv_path.write_text('N\tM\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, ['Col1', 'Col2'], sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )
      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertEqual(sheet.max_row, 3)
      self.assertEqual(sheet.cell(row=3, column=1).value, 'N')

  def test_row_fill_same_day_stays_white(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']
      day = datetime(2026, 7, 12, 10, 0, 0)

      csv_path.write_text('A\tB\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=day,
      )
      csv_path.write_text('C\tD\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=day.replace(hour=15),
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertEqual(sheet.cell(row=2, column=1).fill.fgColor.rgb, ROW_FILL_WHITE.fgColor.rgb)
      self.assertEqual(sheet.cell(row=3, column=1).fill.fgColor.rgb, ROW_FILL_WHITE.fgColor.rgb)

  def test_row_fill_alternates_on_new_day(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']

      csv_path.write_text('A\tB\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY,
        processed_on=datetime(2026, 7, 12, 9, 0, 0),
      )
      csv_path.write_text('C\tD\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY,
        processed_on=datetime(2026, 7, 13, 9, 0, 0),
      )
      csv_path.write_text('E\tF\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY,
        processed_on=datetime(2026, 7, 14, 9, 0, 0),
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertEqual(sheet.cell(row=2, column=1).fill.fgColor.rgb, ROW_FILL_WHITE.fgColor.rgb)
      self.assertEqual(sheet.cell(row=3, column=1).fill.fgColor.rgb, ROW_FILL_PURPLE.fgColor.rgb)
      self.assertEqual(sheet.cell(row=4, column=1).fill.fgColor.rgb, ROW_FILL_WHITE.fgColor.rgb)
      self.assertIn(META_SHEET_NAME, workbook.sheetnames)
      self.assertEqual(workbook[META_SHEET_NAME].sheet_state, 'hidden')

  def test_color_period_anchor_uses_cutoff_hour(self):
    self.assertEqual(
      color_period_anchor(datetime(2026, 7, 21, 20, 0, 0), 19),
      date(2026, 7, 21),
    )
    self.assertEqual(
      color_period_anchor(datetime(2026, 7, 22, 10, 0, 0), 19),
      date(2026, 7, 21),
    )
    self.assertEqual(
      color_period_anchor(datetime(2026, 7, 22, 19, 0, 0), 19),
      date(2026, 7, 22),
    )

  def test_row_fill_same_period_across_midnight_before_cutoff(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']

      csv_path.write_text('A\tB\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY,
        processed_on=datetime(2026, 7, 21, 20, 0, 0),
        row_color_cutoff_hour=19,
      )
      csv_path.write_text('C\tD\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY,
        processed_on=datetime(2026, 7, 22, 10, 0, 0),
        row_color_cutoff_hour=19,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertEqual(sheet.cell(row=2, column=1).fill.fgColor.rgb, ROW_FILL_WHITE.fgColor.rgb)
      self.assertEqual(sheet.cell(row=3, column=1).fill.fgColor.rgb, ROW_FILL_WHITE.fgColor.rgb)

  def test_row_fill_switches_after_cutoff_hour(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']

      csv_path.write_text('A\tB\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY,
        processed_on=datetime(2026, 7, 21, 18, 0, 0),
        row_color_cutoff_hour=19,
      )
      csv_path.write_text('C\tD\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY,
        processed_on=datetime(2026, 7, 21, 20, 0, 0),
        row_color_cutoff_hour=19,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertEqual(sheet.cell(row=2, column=1).fill.fgColor.rgb, ROW_FILL_WHITE.fgColor.rgb)
      self.assertEqual(sheet.cell(row=3, column=1).fill.fgColor.rgb, ROW_FILL_PURPLE.fgColor.rgb)

  def test_header_style_and_column_width(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      long_value = 'descricao_bem_longa_para_testar_largura_da_coluna'
      csv_path.write_text(f'ID\t{long_value}\n', encoding='utf-8')
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Código', 'Descrição detalhada']
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      header = sheet.cell(row=1, column=2)
      data = sheet.cell(row=2, column=2)

      self.assertEqual(header.fill.fgColor.rgb, HEADER_FILL.fgColor.rgb)
      self.assertTrue(header.font.bold)
      self.assertEqual(header.font.color.rgb, HEADER_FONT.color.rgb)
      self.assertEqual(header.alignment.horizontal, CENTER_ALIGNMENT.horizontal)
      self.assertEqual(data.alignment.horizontal, CENTER_ALIGNMENT.horizontal)
      self.assertEqual(header.number_format, TEXT_NUMBER_FORMAT)
      self.assertEqual(data.number_format, TEXT_NUMBER_FORMAT)
      self.assertGreaterEqual(sheet.column_dimensions['B'].width, len(long_value) + 4)

  def test_sheet_hides_gridlines_and_applies_borders(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('A\tB\n', encoding='utf-8')
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertFalse(sheet.sheet_view.showGridLines)
      header = sheet.cell(row=1, column=1)
      data = sheet.cell(row=2, column=1)
      self.assertEqual(header.border.left.style, CELL_BORDER.left.style)
      self.assertEqual(data.border.left.style, CELL_BORDER.left.style)

  def test_numeric_column_writes_number(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('123,45\ttexto\n', encoding='utf-8')
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Valor', 'Descricao']
      append_csv_to_excel(
        csv_path, excel_path, headers, column_types=['number', 'text'],
        sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertEqual(sheet.cell(row=2, column=1).value, 123.45)
      self.assertEqual(sheet.cell(row=2, column=1).number_format, GENERAL_NUMBER_FORMAT)
      self.assertEqual(sheet.cell(row=2, column=2).value, 'texto')
      self.assertEqual(sheet.cell(row=2, column=2).number_format, TEXT_NUMBER_FORMAT)

  def test_numeric_column_invalid_value_falls_back_to_text(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('abc\n', encoding='utf-8')
      excel_path = Path(tmp) / 'saida.xlsx'
      append_csv_to_excel(
        csv_path, excel_path, ['Codigo'], column_types=['number'],
        sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertEqual(sheet.cell(row=2, column=1).value, 'abc')
      self.assertEqual(sheet.cell(row=2, column=1).number_format, TEXT_NUMBER_FORMAT)

  def test_date_column_writes_date(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('12/07/2026\ttexto\n', encoding='utf-8')
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Data', 'Descricao']
      append_csv_to_excel(
        csv_path, excel_path, headers, column_types=['date', 'text'],
        sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      cell_value = sheet.cell(row=2, column=1).value
      if isinstance(cell_value, datetime):
        cell_value = cell_value.date()
      self.assertEqual(cell_value, date(2026, 7, 12))
      self.assertEqual(sheet.cell(row=2, column=1).number_format, DATE_NUMBER_FORMAT)

  def test_date_column_invalid_value_falls_back_to_text(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('nao-e-data\n', encoding='utf-8')
      excel_path = Path(tmp) / 'saida.xlsx'
      append_csv_to_excel(
        csv_path, excel_path, ['Data'], column_types=['date'],
        sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertEqual(sheet.cell(row=2, column=1).value, 'nao-e-data')
      self.assertEqual(sheet.cell(row=2, column=1).number_format, TEXT_NUMBER_FORMAT)

  def test_bold_header_gets_extra_column_width(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('x\n', encoding='utf-8')
      excel_path = Path(tmp) / 'saida.xlsx'
      long_header = 'Descricao detalhada do item'
      append_csv_to_excel(
        csv_path, excel_path, [long_header, 'Outro'], sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertGreaterEqual(sheet.column_dimensions['A'].width, len(long_header) * 1.1)

  def test_append_to_existing_workbook_only_styles_new_rows(self):
    with tempfile.TemporaryDirectory() as tmp:
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']
      csv_path = Path(tmp) / 'dados.csv'
      csv_path.write_text('A\tB\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      sheet.cell(row=5000, column=1, value='')
      sheet.cell(row=5000, column=2, value='')
      workbook.save(excel_path)
      workbook.close()

      csv_path.write_text('C\tD\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      sheet = workbook[self.SHEET_JULY]
      self.assertEqual(sheet.cell(row=2, column=1).value, 'A')
      self.assertEqual(sheet.cell(row=3, column=1).value, 'C')

  def test_last_data_row_cached_in_meta(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']

      csv_path.write_text('A\tB\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      csv_path.write_text('C\tD\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
      )

      workbook = load_workbook(excel_path)
      meta = workbook[META_SHEET_NAME]
      self.assertEqual(meta.cell(row=2, column=5).value, 3)
      self.assertEqual(workbook[self.SHEET_JULY].cell(row=3, column=1).value, 'C')

  def test_duplicate_row_allowed_when_check_disabled(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']

      csv_path.write_text('A\tB\n', encoding='utf-8')
      append_csv_to_excel(csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026)

      csv_path.write_text('A\tB\n', encoding='utf-8')
      sheet_name, count = append_csv_to_excel(
        csv_path,
        excel_path,
        headers,
        sheet_name=self.SHEET_JULY,
        processed_on=self.JULY_2026,
        skip_duplicate_row_check=True,
      )
      self.assertEqual(count, 1)
      workbook = load_workbook(excel_path)
      sheet = workbook[sheet_name]
      self.assertEqual(sheet.cell(row=2, column=1).value, 'A')
      self.assertEqual(sheet.cell(row=3, column=1).value, 'A')

  def test_duplicate_row_in_existing_sheet_raises(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']

      csv_path.write_text('A\tB\n', encoding='utf-8')
      append_csv_to_excel(csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026)

      csv_path.write_text('A\tB\n', encoding='utf-8')
      with self.assertRaises(Exception) as ctx:
        append_csv_to_excel(
          csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
        )
      self.assertIn('duplicada', str(ctx.exception).lower())

  def test_duplicate_row_within_same_file_raises(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']
      csv_path.write_text('A\tB\nA\tB\n', encoding='utf-8')

      with self.assertRaises(Exception) as ctx:
        append_csv_to_excel(
          csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
        )
      self.assertIn('mesmo arquivo', str(ctx.exception).lower())

  def test_duplicate_row_detected_when_meta_last_data_row_is_stale(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Col1', 'Col2']

      csv_path.write_text('A\tB\n', encoding='utf-8')
      append_csv_to_excel(csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026)
      csv_path.write_text('C\tD\n', encoding='utf-8')
      append_csv_to_excel(csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026)

      workbook = load_workbook(excel_path)
      meta = workbook[META_SHEET_NAME]
      meta.cell(row=2, column=5, value=1)
      meta.cell(row=2, column=4, value=1)
      workbook.save(excel_path)

      csv_path.write_text('C\tD\n', encoding='utf-8')
      with self.assertRaises(Exception) as ctx:
        append_csv_to_excel(
          csv_path, excel_path, headers, sheet_name=self.SHEET_JULY, processed_on=self.JULY_2026,
        )
      self.assertIn('duplicada', str(ctx.exception).lower())

  def test_duplicate_row_detected_for_number_columns(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Valor']
      csv_path.write_text('1234,56\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path,
        excel_path,
        headers,
        column_types=['number'],
        sheet_name=self.SHEET_JULY,
        processed_on=self.JULY_2026,
      )

      csv_path.write_text('1234,56\n', encoding='utf-8')
      with self.assertRaises(Exception) as ctx:
        append_csv_to_excel(
          csv_path,
          excel_path,
          headers,
          column_types=['number'],
          sheet_name=self.SHEET_JULY,
          processed_on=self.JULY_2026,
        )
      self.assertIn('duplicada', str(ctx.exception).lower())

  def test_duplicate_row_detected_for_date_columns(self):
    with tempfile.TemporaryDirectory() as tmp:
      csv_path = Path(tmp) / 'dados.csv'
      excel_path = Path(tmp) / 'saida.xlsx'
      headers = ['Data']
      csv_path.write_text('10/07/2026\n', encoding='utf-8')
      append_csv_to_excel(
        csv_path,
        excel_path,
        headers,
        column_types=['date'],
        sheet_name=self.SHEET_JULY,
        processed_on=self.JULY_2026,
      )

      csv_path.write_text('10/07/2026\n', encoding='utf-8')
      with self.assertRaises(Exception) as ctx:
        append_csv_to_excel(
          csv_path,
          excel_path,
          headers,
          column_types=['date'],
          sheet_name=self.SHEET_JULY,
          processed_on=self.JULY_2026,
        )
      self.assertIn('duplicada', str(ctx.exception).lower())


class FlowTests(unittest.TestCase):
  def test_roundtrip(self):
    flow = Flow(
      name='Teste',
      source_filename='a.csv',
      excel_directory='/tmp',
      excel_filename='out.xlsx',
      headers=['A', 'B', 'C'],
      column_types=['text', 'number', 'date'],
      column_duplicate_checks=[True, False, False],
    )
    restored = Flow.from_dict(flow.to_dict())
    self.assertEqual(restored.name, flow.name)
    self.assertEqual(restored.headers, flow.headers)
    self.assertEqual(restored.normalized_column_types(), ['text', 'number', 'date'])
    self.assertEqual(restored.normalized_column_duplicate_checks(), [True, False, False])


if __name__ == '__main__':
  unittest.main()
