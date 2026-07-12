from __future__ import annotations

import csv
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.models.schema import normalize_column_type
from app.services.excel_crypto import (
  close_workbook,
  create_empty_workbook,
  load_workbook_from_path,
  save_workbook_to_path,
)

META_SHEET_NAME = '__AEF__'
ROW_FILL_WHITE = PatternFill(fill_type='solid', fgColor='FFFFFF')
ROW_FILL_PURPLE = PatternFill(fill_type='solid', fgColor='E9D5FF')
ROW_FILLS = (ROW_FILL_WHITE, ROW_FILL_PURPLE)
HEADER_FILL = PatternFill(fill_type='solid', fgColor='6D28D9')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
DUPLICATE_FILL = PatternFill(fill_type='solid', fgColor='FECACA')
DATA_FONT = Font(name='Calibri', size=8)
DUPLICATE_DATA_FONT = Font(name='Calibri', size=8, color='B91C1C')
TEXT_NUMBER_FORMAT = '@'
GENERAL_NUMBER_FORMAT = 'General'
DATE_NUMBER_FORMAT = 'DD/MM/YYYY'
MIN_COLUMN_WIDTH = 8
MAX_COLUMN_WIDTH = 120
COLUMN_WIDTH_PADDING = 4
BOLD_WIDTH_FACTOR = 1.15
DATA_START_ROW = 2
CELL_BORDER = Border(
  left=Side(style='thin', color='C4B5FD'),
  right=Side(style='thin', color='C4B5FD'),
  top=Side(style='thin', color='C4B5FD'),
  bottom=Side(style='thin', color='C4B5FD'),
)
NO_BORDER = Border()


class DuplicateRowError(ValueError):
  def __init__(
    self,
    incoming_row_number: int,
    matching_row_number: int,
    row_values: List[str],
    *,
    matching_source: str = 'planilha',
  ):
    self.incoming_row_number = incoming_row_number
    self.matching_row_number = matching_row_number
    self.row_values = row_values
    self.matching_source = matching_source
    values = ' | '.join(row_values)
    super().__init__(
      f'Linha {incoming_row_number} duplicada (igual à linha {matching_row_number} '
      f'da {matching_source}): {values}'
    )

MONTHS_PT = [
  'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
  'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro',
]


def current_month_sheet_name(when: datetime | None = None) -> str:
  when = when or datetime.now()
  return MONTHS_PT[when.month - 1]


def _resolve_sheet_name(workbook, sheet_name: str) -> str:
  if sheet_name in workbook.sheetnames:
    return sheet_name
  target = sheet_name.lower()
  for existing_name in workbook.sheetnames:
    if existing_name.lower() == target:
      return existing_name
  return sheet_name


def excel_year_directory(directory: str, when: datetime | None = None) -> Path:
  when = when or datetime.now()
  return Path(directory) / str(when.year)


def excel_full_path(directory: str, filename: str, when: datetime | None = None) -> Path:
  name = filename if filename.lower().endswith('.xlsx') else f'{filename}.xlsx'
  return excel_year_directory(directory, when) / name


CSV_ENCODINGS = ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1')


def read_tab_csv(path: Path) -> List[List[str]]:
  raw = path.read_bytes()
  last_error: UnicodeDecodeError | None = None
  text: str | None = None

  for encoding in CSV_ENCODINGS:
    try:
      text = raw.decode(encoding)
      break
    except UnicodeDecodeError as exc:
      last_error = exc

  if text is None:
    raise ValueError(
      f'Não foi possível decodificar {path.name}. Tentativas: {", ".join(CSV_ENCODINGS)}.'
    ) from last_error

  rows: List[List[str]] = []
  for row in csv.reader(text.splitlines(), delimiter='\t'):
    if not row or all(not cell.strip() for cell in row):
      continue
    rows.append([cell.strip() for cell in row])
  return rows


def _sheet_is_empty(sheet: Worksheet) -> bool:
  if sheet.max_row == 1 and sheet.max_column == 1:
    value = sheet.cell(row=1, column=1).value
    return value is None or str(value).strip() == ''
  return sheet.max_row == 0


def _write_text_cell(sheet: Worksheet, row_index: int, col_index: int, value) -> None:
  cell = sheet.cell(row=row_index, column=col_index, value='' if value is None else str(value))
  cell.number_format = TEXT_NUMBER_FORMAT
  cell.font = DATA_FONT
  cell.alignment = CENTER_ALIGNMENT


def _parse_numeric_value(value: str) -> int | float | None:
  text = value.strip()
  if not text:
    return None
  normalized = text
  if ',' in normalized and '.' in normalized:
    normalized = normalized.replace('.', '').replace(',', '.')
  elif ',' in normalized:
    normalized = normalized.replace(',', '.')
  try:
    number = float(normalized)
  except ValueError:
    return None
  if number.is_integer():
    return int(number)
  return number


def _parse_date_value(value: str) -> date | None:
  text = value.strip()
  if not text:
    return None
  for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d.%m.%Y'):
    try:
      return datetime.strptime(text, fmt).date()
    except ValueError:
      continue
  return None


def _write_data_cell(
  sheet: Worksheet,
  row_index: int,
  col_index: int,
  value,
  *,
  column_type: str = 'text',
) -> None:
  text = '' if value is None else str(value).strip()
  if column_type == 'number':
    number = _parse_numeric_value(text)
    if number is not None:
      cell = sheet.cell(row=row_index, column=col_index, value=number)
      cell.number_format = GENERAL_NUMBER_FORMAT
      cell.font = DATA_FONT
      cell.alignment = CENTER_ALIGNMENT
      return
  if column_type == 'date':
    parsed_date = _parse_date_value(text)
    if parsed_date is not None:
      cell = sheet.cell(row=row_index, column=col_index, value=parsed_date)
      cell.number_format = DATE_NUMBER_FORMAT
      cell.font = DATA_FONT
      cell.alignment = CENTER_ALIGNMENT
      return
  _write_text_cell(sheet, row_index, col_index, text)


def _ensure_headers(sheet: Worksheet, headers: List[str]) -> None:
  if not headers:
    return
  if _sheet_is_empty(sheet):
    for col_index, header in enumerate(headers, start=1):
      _write_text_cell(sheet, 1, col_index, header)
    return
  existing = [str(sheet.cell(row=1, column=col).value or '').strip() for col in range(1, sheet.max_column + 1)]
  if not any(existing):
    for col_index, header in enumerate(headers, start=1):
      _write_text_cell(sheet, 1, col_index, header)


def _last_data_row(sheet: Worksheet, column_count: int) -> int:
  last_row = 1
  for row_index in range(DATA_START_ROW, sheet.max_row + 1):
    values = _row_values_from_sheet(sheet, row_index, column_count)
    if not _row_is_blank(values):
      last_row = row_index
  return last_row


def _next_data_row(sheet: Worksheet, column_count: int) -> int:
  if _sheet_is_empty(sheet):
    return 2 if sheet.max_row >= 1 and sheet.cell(row=1, column=1).value else 1
  return _last_data_row(sheet, column_count) + 1


def _ensure_meta_sheet(workbook) -> Worksheet:
  if META_SHEET_NAME in workbook.sheetnames:
    return workbook[META_SHEET_NAME]
  sheet = workbook.create_sheet(title=META_SHEET_NAME)
  sheet.sheet_state = 'hidden'
  sheet.cell(row=1, column=1, value='sheet_name')
  sheet.cell(row=1, column=2, value='last_date')
  sheet.cell(row=1, column=3, value='color_index')
  sheet.cell(row=1, column=4, value='layout_done')
  sheet.cell(row=1, column=5, value='last_data_row')
  return sheet


def _ensure_meta_headers(meta_sheet: Worksheet) -> None:
  for column_index, header in (
    (4, 'layout_done'),
    (5, 'last_data_row'),
  ):
    if meta_sheet.cell(row=1, column=column_index).value in (None, ''):
      meta_sheet.cell(row=1, column=column_index, value=header)


def _parse_meta_date(value) -> date | None:
  if value is None:
    return None
  if isinstance(value, datetime):
    return value.date()
  if isinstance(value, date):
    return value
  text = str(value).strip()
  if not text:
    return None
  return date.fromisoformat(text)


def _parse_layout_done(value) -> bool:
  if value is None:
    return False
  if isinstance(value, bool):
    return value
  if isinstance(value, (int, float)):
    return int(value) == 1
  return str(value).strip().lower() in {'1', 'true', 'sim', 'yes'}


def _parse_last_data_row(value) -> int | None:
  if value is None or value == '':
    return None
  try:
    parsed = int(value)
  except (TypeError, ValueError):
    return None
  return parsed if parsed >= 1 else None


def _read_sheet_meta(meta_sheet: Worksheet, sheet_name: str) -> Tuple[date | None, int, bool, int | None]:
  for row_index in range(2, meta_sheet.max_row + 1):
    if str(meta_sheet.cell(row=row_index, column=1).value or '').strip() != sheet_name:
      continue
    last_date = _parse_meta_date(meta_sheet.cell(row=row_index, column=2).value)
    color_index = int(meta_sheet.cell(row=row_index, column=3).value or 0)
    layout_done = _parse_layout_done(meta_sheet.cell(row=row_index, column=4).value)
    last_data_row = _parse_last_data_row(meta_sheet.cell(row=row_index, column=5).value)
    return last_date, 0 if color_index not in (0, 1) else color_index, layout_done, last_data_row
  return None, 0, False, None


def _read_day_color_state(meta_sheet: Worksheet, sheet_name: str) -> Tuple[date | None, int]:
  last_date, color_index, _layout_done, _last_data_row = _read_sheet_meta(meta_sheet, sheet_name)
  return last_date, color_index


def _write_sheet_meta(
  meta_sheet: Worksheet,
  sheet_name: str,
  process_date: date,
  color_index: int,
  *,
  layout_done: bool,
  last_data_row: int | None = None,
) -> None:
  for row_index in range(2, meta_sheet.max_row + 1):
    if str(meta_sheet.cell(row=row_index, column=1).value or '').strip() == sheet_name:
      meta_sheet.cell(row=row_index, column=2, value=process_date.isoformat())
      meta_sheet.cell(row=row_index, column=3, value=color_index)
      meta_sheet.cell(row=row_index, column=4, value=1 if layout_done else 0)
      if last_data_row is not None:
        meta_sheet.cell(row=row_index, column=5, value=int(last_data_row))
      return
  next_row = max(meta_sheet.max_row, 1) + 1
  meta_sheet.cell(row=next_row, column=1, value=sheet_name)
  meta_sheet.cell(row=next_row, column=2, value=process_date.isoformat())
  meta_sheet.cell(row=next_row, column=3, value=color_index)
  meta_sheet.cell(row=next_row, column=4, value=1 if layout_done else 0)
  if last_data_row is not None:
    meta_sheet.cell(row=next_row, column=5, value=int(last_data_row))


def _write_day_color_state(meta_sheet: Worksheet, sheet_name: str, process_date: date, color_index: int) -> None:
  _last_date, _color_index, layout_done, last_data_row = _read_sheet_meta(meta_sheet, sheet_name)
  _write_sheet_meta(
    meta_sheet,
    sheet_name,
    process_date,
    color_index,
    layout_done=layout_done,
    last_data_row=last_data_row,
  )


def _resolve_start_row(
  sheet: Worksheet,
  column_count: int,
  cached_last_data_row: int | None,
) -> Tuple[int, int]:
  if _sheet_is_empty(sheet):
    if sheet.max_row >= 1 and sheet.cell(row=1, column=1).value:
      return DATA_START_ROW, 1
    return 1, 1

  if cached_last_data_row is not None and cached_last_data_row >= 1:
    return max(cached_last_data_row + 1, DATA_START_ROW), cached_last_data_row

  resolved_last = _last_data_row(sheet, column_count)
  return max(resolved_last + 1, DATA_START_ROW), resolved_last


def _resolve_row_color_index(last_date: date | None, current_date: date, current_index: int) -> int:
  if last_date is None:
    return 0
  if last_date == current_date:
    return current_index
  return 1 - current_index


def _row_fill_column_count(sheet: Worksheet, headers: List[str], rows: List[List[str]]) -> int:
  max_cols = sheet.max_column
  if headers:
    max_cols = max(max_cols, len(headers))
  for row in rows:
    max_cols = max(max_cols, len(row))
  return max(max_cols, 1)


def _apply_row_fill(sheet: Worksheet, row_index: int, fill: PatternFill, column_count: int) -> None:
  for col_index in range(1, column_count + 1):
    sheet.cell(row=row_index, column=col_index).fill = fill


def _configure_sheet_view(sheet: Worksheet) -> None:
  sheet.sheet_view.showGridLines = False


def _sheet_bounds(sheet: Worksheet) -> Tuple[int, int, int, int]:
  dimension = sheet.calculate_dimension()
  if not dimension or dimension == 'A1:A1':
    value = sheet.cell(row=1, column=1).value
    if value is None or str(value).strip() == '':
      return 1, 1, 1, 1
  min_col, min_row, max_col, max_row = range_boundaries(dimension)
  return min_row, max_row, min_col, max_col


def _clear_borders_in_range(
  sheet: Worksheet,
  *,
  min_row: int,
  max_row: int,
  min_col: int,
  max_col: int,
) -> None:
  for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
    for cell in row:
      cell.border = NO_BORDER


def _migrate_existing_sheet_layout(sheet: Worksheet, column_count: int) -> None:
  _configure_sheet_view(sheet)
  min_row, max_row, min_col, max_col = _sheet_bounds(sheet)
  max_col = max(max_col, column_count)
  _clear_borders_in_range(
    sheet,
    min_row=min_row,
    max_row=max_row,
    min_col=min_col,
    max_col=max_col,
  )
  _apply_header_style(sheet, column_count)
  data_end = max(_last_data_row(sheet, column_count), 1)
  _apply_data_row_fonts(sheet, column_count, end_row=data_end)
  _apply_table_borders(sheet, column_count, start_row=1, end_row=data_end)


def _estimate_text_width(text: str, *, bold: bool = False) -> float:
  width = 0.0
  for char in text:
    width += 2.0 if ord(char) > 127 else 1.0
  if bold:
    width *= BOLD_WIDTH_FACTOR
  return width


def _apply_table_borders(
  sheet: Worksheet,
  column_count: int,
  *,
  start_row: int = 1,
  end_row: int | None = None,
) -> None:
  end_row = end_row if end_row is not None else sheet.max_row
  for row_index in range(start_row, end_row + 1):
    if row_index >= DATA_START_ROW:
      values = _row_values_from_sheet(sheet, row_index, column_count)
      if _row_is_blank(values):
        continue
    for col_index in range(1, column_count + 1):
      if row_index == 1:
        cell = sheet.cell(row=row_index, column=col_index)
        if cell.value is None or str(cell.value).strip() == '':
          continue
      sheet.cell(row=row_index, column=col_index).border = CELL_BORDER


def _apply_header_style(sheet: Worksheet, column_count: int) -> None:
  for col_index in range(1, column_count + 1):
    cell = sheet.cell(row=1, column=col_index)
    if cell.value is None or str(cell.value).strip() == '':
      continue
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.number_format = TEXT_NUMBER_FORMAT
    cell.alignment = CENTER_ALIGNMENT


def _apply_data_row_fonts(
  sheet: Worksheet,
  column_count: int,
  *,
  start_row: int = DATA_START_ROW,
  end_row: int,
) -> None:
  for row_index in range(start_row, end_row + 1):
    if _row_is_blank(_row_values_from_sheet(sheet, row_index, column_count)):
      continue
    for col_index in range(1, column_count + 1):
      sheet.cell(row=row_index, column=col_index).font = DATA_FONT


def _column_compare_value(value, *, column_type: str) -> str:
  if value is None:
    return ''
  if column_type == 'number' and isinstance(value, (int, float)):
    return str(value)
  if column_type == 'date':
    if isinstance(value, datetime):
      return value.date().isoformat()
    if isinstance(value, date):
      return value.isoformat()
  return str(value).strip()


def _reference_fill_column(highlight_columns: set[int], column_count: int) -> int:
  for col_index in range(1, column_count + 1):
    if col_index not in highlight_columns:
      return col_index
  return 1


def _apply_duplicate_column_highlights(
  sheet: Worksheet,
  column_count: int,
  *,
  column_types: List[str],
  highlight_columns: List[bool],
  end_row: int,
) -> None:
  enabled_columns = {
    col_index
    for col_index, enabled in enumerate(highlight_columns[:column_count], start=1)
    if enabled
  }
  if not enabled_columns:
    return

  data_end = max(end_row, _last_data_row(sheet, column_count))
  reference_col = _reference_fill_column(enabled_columns, column_count)

  for col_index in enabled_columns:
    column_type = column_types[col_index - 1] if col_index <= len(column_types) else 'text'
    values_by_row: dict[int, str] = {}
    for row_index in range(DATA_START_ROW, data_end + 1):
      if _row_is_blank(_row_values_from_sheet(sheet, row_index, column_count)):
        continue
      value = _column_compare_value(
        sheet.cell(row=row_index, column=col_index).value,
        column_type=column_type,
      )
      if value:
        values_by_row[row_index] = value

    counts: dict[str, int] = {}
    for value in values_by_row.values():
      counts[value] = counts.get(value, 0) + 1

    duplicated_values = {value for value, count in counts.items() if count > 1}
    for row_index, value in values_by_row.items():
      cell = sheet.cell(row=row_index, column=col_index)
      if value in duplicated_values:
        cell.fill = DUPLICATE_FILL
        cell.font = DUPLICATE_DATA_FONT
      else:
        reference_fill = sheet.cell(row=row_index, column=reference_col).fill
        if reference_fill and reference_fill.fill_type:
          cell.fill = copy(reference_fill)
        else:
          cell.fill = ROW_FILL_WHITE
      cell.alignment = CENTER_ALIGNMENT


def _apply_duplicate_column_highlights_incremental(
  sheet: Worksheet,
  column_count: int,
  *,
  column_types: List[str],
  highlight_columns: List[bool],
  new_start_row: int,
  new_end_row: int,
  existing_last_row: int | None = None,
) -> None:
  if new_start_row > new_end_row:
    return

  enabled_columns = {
    col_index
    for col_index, enabled in enumerate(highlight_columns[:column_count], start=1)
    if enabled
  }
  if not enabled_columns:
    return

  data_end = max(new_end_row, existing_last_row or 0)
  if data_end < DATA_START_ROW:
    return
  reference_col = _reference_fill_column(enabled_columns, column_count)

  for col_index in enabled_columns:
    column_type = column_types[col_index - 1] if col_index <= len(column_types) else 'text'
    tracked_values: set[str] = set()
    for row_index in range(new_start_row, new_end_row + 1):
      if _row_is_blank(_row_values_from_sheet(sheet, row_index, column_count)):
        continue
      value = _column_compare_value(
        sheet.cell(row=row_index, column=col_index).value,
        column_type=column_type,
      )
      if value:
        tracked_values.add(value)
    if not tracked_values:
      continue

    counts = {value: 0 for value in tracked_values}
    rows_by_value: dict[str, list[int]] = {value: [] for value in tracked_values}
    for row_index in range(DATA_START_ROW, data_end + 1):
      if _row_is_blank(_row_values_from_sheet(sheet, row_index, column_count)):
        continue
      value = _column_compare_value(
        sheet.cell(row=row_index, column=col_index).value,
        column_type=column_type,
      )
      if value not in tracked_values:
        continue
      counts[value] += 1
      rows_by_value[value].append(row_index)

    duplicated_values = {value for value, count in counts.items() if count > 1}
    for value, row_indexes in rows_by_value.items():
      for row_index in row_indexes:
        cell = sheet.cell(row=row_index, column=col_index)
        if value in duplicated_values:
          cell.fill = DUPLICATE_FILL
          cell.font = DUPLICATE_DATA_FONT
        else:
          reference_fill = sheet.cell(row=row_index, column=reference_col).fill
          if reference_fill and reference_fill.fill_type:
            cell.fill = copy(reference_fill)
          else:
            cell.fill = ROW_FILL_WHITE
          cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGNMENT


def _apply_full_sheet_formatting(
  sheet: Worksheet,
  column_count: int,
  *,
  column_types: List[str],
  highlight_columns: List[bool],
  end_row: int,
) -> None:
  _configure_sheet_view(sheet)
  _apply_header_style(sheet, column_count)
  _apply_table_borders(sheet, column_count, start_row=1, end_row=end_row)
  _autofit_column_widths(sheet, column_count, start_row=1, end_row=end_row)
  _apply_center_alignment(sheet, column_count, start_row=1, end_row=end_row)
  _apply_data_row_fonts(sheet, column_count, end_row=end_row)
  _apply_duplicate_column_highlights(
    sheet,
    column_count,
    column_types=column_types,
    highlight_columns=highlight_columns,
    end_row=end_row,
  )


def _apply_new_rows_formatting(
  sheet: Worksheet,
  column_count: int,
  *,
  column_types: List[str],
  highlight_columns: List[bool],
  start_row: int,
  end_row: int,
  existing_last_row: int | None = None,
) -> None:
  _apply_table_borders(sheet, column_count, start_row=start_row, end_row=end_row)
  _autofit_column_widths(sheet, column_count, start_row=start_row, end_row=end_row)
  _apply_center_alignment(sheet, column_count, start_row=start_row, end_row=end_row)
  _apply_data_row_fonts(sheet, column_count, start_row=start_row, end_row=end_row)
  _apply_duplicate_column_highlights_incremental(
    sheet,
    column_count,
    column_types=column_types,
    highlight_columns=highlight_columns,
    new_start_row=start_row,
    new_end_row=end_row,
    existing_last_row=existing_last_row,
  )


def _apply_center_alignment(
  sheet: Worksheet,
  column_count: int,
  *,
  start_row: int = 1,
  end_row: int,
) -> None:
  for row_index in range(start_row, end_row + 1):
    for col_index in range(1, column_count + 1):
      sheet.cell(row=row_index, column=col_index).alignment = CENTER_ALIGNMENT


def _normalize_row_values(row: List[str], column_count: int) -> Tuple[str, ...]:
  values: List[str] = []
  for col_index in range(column_count):
    if col_index < len(row):
      values.append(str(row[col_index]).strip())
    else:
      values.append('')
  return tuple(values)


def _row_values_from_sheet(sheet: Worksheet, row_index: int, column_count: int) -> Tuple[str, ...]:
  values: List[str] = []
  for col_index in range(1, column_count + 1):
    values.append(str(sheet.cell(row=row_index, column=col_index).value or '').strip())
  return tuple(values)


def _row_is_blank(values: Tuple[str, ...]) -> bool:
  return not any(values)


def _iter_sheet_row_tuples(
  sheet: Worksheet,
  column_count: int,
  *,
  start_row: int,
  end_row: int,
):
  for row_index, row_values in enumerate(
    sheet.iter_rows(
      min_row=start_row,
      max_row=end_row,
      min_col=1,
      max_col=column_count,
      values_only=True,
    ),
    start=start_row,
  ):
    values = tuple(str(value or '').strip() for value in row_values)
    if _row_is_blank(values):
      continue
    yield row_index, values


def _validate_no_duplicate_rows(
  sheet: Worksheet,
  new_rows: List[List[str]],
  column_count: int,
  *,
  start_row: int,
  existing_last_row: int | None = None,
) -> None:
  known_rows: dict[Tuple[str, ...], Tuple[int, str]] = {}
  scan_end = existing_last_row if existing_last_row is not None else _last_data_row(sheet, column_count)
  if scan_end >= DATA_START_ROW:
    for row_index, values in _iter_sheet_row_tuples(
      sheet,
      column_count,
      start_row=DATA_START_ROW,
      end_row=scan_end,
    ):
      known_rows[values] = (row_index, 'planilha')

  for offset, row in enumerate(new_rows):
    values = _normalize_row_values(row, column_count)
    if _row_is_blank(values):
      continue
    incoming_row_number = offset + 1
    if values in known_rows:
      matching_row_number, matching_source = known_rows[values]
      raise DuplicateRowError(
        incoming_row_number,
        matching_row_number,
        list(values),
        matching_source=matching_source,
      )
    known_rows[values] = (incoming_row_number, 'mesmo arquivo')


def _autofit_column_widths(
  sheet: Worksheet,
  column_count: int,
  *,
  start_row: int = 1,
  end_row: int | None = None,
) -> None:
  end_row = end_row if end_row is not None else sheet.max_row
  for col_index in range(1, column_count + 1):
    max_width = float(MIN_COLUMN_WIDTH)
    for row_index in range(start_row, end_row + 1):
      value = sheet.cell(row=row_index, column=col_index).value
      if value is None:
        continue
      text_width = _estimate_text_width(str(value), bold=row_index == 1)
      max_width = max(max_width, text_width)
    width = min(max(max_width + COLUMN_WIDTH_PADDING, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
    col_letter = get_column_letter(col_index)
    current_width = sheet.column_dimensions[col_letter].width or 0
    if width > current_width:
      sheet.column_dimensions[col_letter].width = width


def _text_column_sqref(column_count: int, column_types: List[str]) -> str:
  ranges: List[str] = []
  for col_index in range(1, column_count + 1):
    column_type = column_types[col_index - 1] if col_index <= len(column_types) else 'text'
    if column_type != 'text':
      continue
    letter = get_column_letter(col_index)
    ranges.append(f'{letter}1:{letter}1048576')
  return ' '.join(ranges)


def append_csv_to_excel(
  csv_path: Path,
  excel_path: Path,
  headers: List[str],
  *,
  column_types: List[str] | None = None,
  column_duplicate_checks: List[bool] | None = None,
  sheet_name: str | None = None,
  excel_password: Optional[str] = None,
  processed_on: datetime | None = None,
  skip_duplicate_row_check: bool = False,
  extra_csv_paths: List[Path] | None = None,
) -> Tuple[str, int]:
  sheet_name = sheet_name or current_month_sheet_name()
  process_date = (processed_on or datetime.now()).date()
  csv_paths = [csv_path, *(extra_csv_paths or [])]
  rows: List[List[str]] = []
  for path in csv_paths:
    rows.extend(read_tab_csv(path))
  if not rows:
    names = ', '.join(path.name for path in csv_paths)
    raise ValueError(f'Arquivo(s) CSV vazio(s) ou sem linhas válidas: {names}.')

  normalized_types = [normalize_column_type(item) for item in (column_types or [])]
  while len(normalized_types) < len(headers):
    normalized_types.append('text')
  normalized_types = normalized_types[:len(headers)]

  duplicate_checks = [bool(item) for item in (column_duplicate_checks or [])]
  while len(duplicate_checks) < len(headers):
    duplicate_checks.append(False)
  duplicate_checks = duplicate_checks[:len(headers)]

  excel_path.parent.mkdir(parents=True, exist_ok=True)
  workbook = None
  try:
    if excel_path.is_file():
      workbook = load_workbook_from_path(excel_path, password=excel_password or None)
    else:
      workbook = create_empty_workbook()

    sheet_name = _resolve_sheet_name(workbook, sheet_name)
    sheet_already_existed = sheet_name in workbook.sheetnames
    if sheet_already_existed:
      sheet = workbook[sheet_name]
    else:
      sheet = workbook.create_sheet(title=sheet_name)

    _ensure_headers(sheet, headers)
    fill_column_count = _row_fill_column_count(sheet, headers, rows)

    meta_sheet = _ensure_meta_sheet(workbook)
    _ensure_meta_headers(meta_sheet)
    last_date, color_index, layout_done, cached_last_data_row = _read_sheet_meta(meta_sheet, sheet_name)
    if sheet_already_existed and not layout_done:
      _migrate_existing_sheet_layout(sheet, fill_column_count)
      layout_done = True
      if cached_last_data_row is None:
        cached_last_data_row = _last_data_row(sheet, fill_column_count)

    start_row, existing_last_data_row = _resolve_start_row(
      sheet,
      fill_column_count,
      cached_last_data_row,
    )
    if start_row == 1 and headers:
      start_row = DATA_START_ROW
      existing_last_data_row = max(existing_last_data_row, 1)

    color_index = _resolve_row_color_index(last_date, process_date, color_index)
    row_fill = ROW_FILLS[color_index]

    if not skip_duplicate_row_check:
      _validate_no_duplicate_rows(
        sheet,
        rows,
        fill_column_count,
        start_row=start_row,
        existing_last_row=existing_last_data_row,
      )

    for offset, row in enumerate(rows):
      target_row = start_row + offset
      for col_index, value in enumerate(row, start=1):
        column_type = normalized_types[col_index - 1] if col_index <= len(normalized_types) else 'text'
        _write_data_cell(
          sheet,
          target_row,
          col_index,
          value,
          column_type=column_type,
        )
      _apply_row_fill(sheet, target_row, row_fill, fill_column_count)

    last_written_row = start_row + len(rows) - 1 if rows else start_row - 1
    updated_last_data_row = max(existing_last_data_row, last_written_row)

    if not layout_done:
      _apply_full_sheet_formatting(
        sheet,
        fill_column_count,
        column_types=normalized_types,
        highlight_columns=duplicate_checks,
        end_row=updated_last_data_row,
      )
      layout_done = True
    elif rows:
      _apply_new_rows_formatting(
        sheet,
        fill_column_count,
        column_types=normalized_types,
        highlight_columns=duplicate_checks,
        start_row=start_row,
        end_row=last_written_row,
        existing_last_row=existing_last_data_row,
      )

    _write_sheet_meta(
      meta_sheet,
      sheet_name,
      process_date,
      color_index,
      layout_done=layout_done,
      last_data_row=updated_last_data_row,
    )

    ignored_text_sqref = _text_column_sqref(fill_column_count, normalized_types)
    ignored_text_sqref_by_sheet = {sheet_name: ignored_text_sqref} if ignored_text_sqref else None

    save_workbook_to_path(
      workbook,
      excel_path,
      password=excel_password or None,
      ignored_text_sqref_by_sheet=ignored_text_sqref_by_sheet,
    )
    return sheet_name, len(rows)
  finally:
    close_workbook(workbook)
