from __future__ import annotations

import io
import os
import zipfile
from pathlib import Path
from typing import Optional
from xml.etree import ElementTree as ET

import msoffcrypto
from msoffcrypto.exceptions import InvalidKeyError
from openpyxl import Workbook, load_workbook
from openpyxl.utils.protection import hash_password
from openpyxl.workbook.workbook import Workbook as WorkbookType

MAIN_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
REL_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'


class ExcelPasswordError(ValueError):
  pass


class ExcelFileError(ValueError):
  pass


def format_excel_error(exc: BaseException) -> str:
  if isinstance(exc, ExcelPasswordError):
    return str(exc)
  if isinstance(exc, ExcelFileError):
    return str(exc)
  if isinstance(exc, OSError):
    winerror = getattr(exc, 'winerror', None)
    if winerror == 32:
      return 'O arquivo Excel está aberto em outro programa. Feche-o e tente novamente.'
    if winerror in {5, 13}:
      return 'Sem permissão para acessar o arquivo Excel.'
  text = str(exc).strip()
  if not text or 'returned a result with an exception set' in text:
    cause = exc.__cause__ or exc.__context__
    if cause and cause is not exc:
      return format_excel_error(cause)
    return (
      'Erro ao acessar o arquivo Excel. Verifique se ele não está aberto no Excel '
      'e se foi salvo corretamente como .xlsx.'
    )
  return text


def close_workbook(workbook: WorkbookType | None) -> None:
  if workbook is None:
    return
  try:
    workbook.close()
  except OSError:
    pass


def is_encrypted_excel(path: Path) -> bool:
  if not path.is_file():
    return False
  with open(path, 'rb') as handle:
    office_file = msoffcrypto.OfficeFile(handle)
    return bool(office_file.is_encrypted())


def _patch_workbook_password_xml(workbook_xml: bytes, password: str) -> bytes:
  pwd_hash = hash_password(password)
  root = ET.fromstring(workbook_xml)

  for child in list(root):
    if child.tag == f'{{{MAIN_NS}}}fileSharing':
      root.remove(child)

  file_sharing = ET.Element(f'{{{MAIN_NS}}}fileSharing')
  file_sharing.set('readOnlyRecommended', '1')
  file_sharing.set('reservationPassword', pwd_hash)
  file_sharing.set('userName', 'AutomaticExcelFill')
  root.insert(0, file_sharing)

  ET.register_namespace('', MAIN_NS)
  ET.register_namespace('r', REL_NS)
  return ET.tostring(root, encoding='UTF-8', xml_declaration=True)


def _inject_windows_modify_password(path: Path, password: str) -> None:
  """Emula Excel: Salvar como → Ferramentas → Opções gerais → Senha para modificar."""
  _postprocess_saved_xlsx(path, password=password)


def _resolve_ignored_sheet_files(
  workbook: WorkbookType,
  sqref_by_sheet: dict[str, str],
) -> dict[str, str]:
  sheet_files: dict[str, str] = {}
  for worksheet in workbook.worksheets:
    sqref = (sqref_by_sheet.get(worksheet.title) or '').strip()
    if not sqref or worksheet._id is None:
      continue
    sheet_files[f'xl/worksheets/sheet{worksheet._id}.xml'] = sqref
  return sheet_files


def _postprocess_saved_xlsx(
  path: Path,
  *,
  workbook: WorkbookType | None = None,
  ignored_text_sqref_by_sheet: dict[str, str] | None = None,
  password: str | None = None,
) -> None:
  sheet_files = _resolve_ignored_sheet_files(workbook, ignored_text_sqref_by_sheet or {}) if workbook else {}
  if not sheet_files and not password:
    return

  buffer = io.BytesIO()
  with zipfile.ZipFile(path, 'r') as zin:
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zout:
      for item in zin.infolist():
        data = zin.read(item.filename)
        if password and item.filename == 'xl/workbook.xml':
          data = _patch_workbook_password_xml(data, password)
        elif item.filename in sheet_files:
          data = _append_ignored_errors_xml(data, sheet_files[item.filename])
        zout.writestr(item, data)
  path.write_bytes(buffer.getvalue())


def read_file_sharing_password_hash(path: Path) -> Optional[str]:
  if not path.is_file():
    return None
  with zipfile.ZipFile(path, 'r') as zin:
    if 'xl/workbook.xml' not in zin.namelist():
      return None
    root = ET.fromstring(zin.read('xl/workbook.xml'))
  for child in root:
    if child.tag == f'{{{MAIN_NS}}}fileSharing':
      return child.get('reservationPassword')
  return None


def load_workbook_from_path(path: Path, password: Optional[str] = None) -> WorkbookType:
  if not path.is_file():
    raise FileNotFoundError(str(path))

  try:
    raw = path.read_bytes()
  except OSError as exc:
    raise ExcelFileError(format_excel_error(exc)) from exc

  office_file = msoffcrypto.OfficeFile(io.BytesIO(raw))
  if office_file.is_encrypted():
    if not password:
      raise ExcelPasswordError(
        'Este Excel foi salvo com senha de abertura (formato antigo). '
        'Informe a senha no fluxo ou salve novamente pelo app para usar senha para modificar.',
      )
    try:
      office_file.load_key(password=password)
      decrypted = io.BytesIO()
      office_file.decrypt(decrypted)
    except InvalidKeyError as exc:
      raise ExcelPasswordError('Senha do Excel incorreta.') from exc
    except Exception as exc:
      raise ExcelPasswordError('Não foi possível abrir o Excel com a senha informada.') from exc
    decrypted.seek(0)
    try:
      return load_workbook(decrypted, keep_links=False)
    except Exception as exc:
      raise ExcelFileError(
        f'Não foi possível abrir "{path.name}" após descriptografar. '
        f'Verifique se o arquivo é um .xlsx válido.'
      ) from exc

  try:
    return load_workbook(io.BytesIO(raw), keep_links=False)
  except Exception as exc:
    raise ExcelFileError(
      f'Não foi possível abrir "{path.name}". '
      f'O arquivo pode estar corrompido ou conter recursos que o app ainda não suporta. '
      f'Tente abrir e salvar novamente no Excel como .xlsx.'
    ) from exc


def _append_ignored_errors_xml(xml_bytes: bytes, sqref: str) -> bytes:
  root = ET.fromstring(xml_bytes)
  ignored_errors = None
  for child in root:
    if child.tag == f'{{{MAIN_NS}}}ignoredErrors':
      ignored_errors = child
      break

  if ignored_errors is None:
    ignored_errors = ET.Element(f'{{{MAIN_NS}}}ignoredErrors')
    root.append(ignored_errors)

  for child in ignored_errors:
    if child.tag != f'{{{MAIN_NS}}}ignoredError':
      continue
    if child.get('numberStoredAsText') in {'1', 'true'}:
      existing = (child.get('sqref') or '').strip()
      child.set('sqref', sqref if not existing else f'{existing} {sqref}')
      break
  else:
    ignored_error = ET.Element(f'{{{MAIN_NS}}}ignoredError')
    ignored_error.set('sqref', sqref)
    ignored_error.set('numberStoredAsText', '1')
    ignored_errors.append(ignored_error)

  ET.register_namespace('', MAIN_NS)
  return ET.tostring(root, encoding='UTF-8', xml_declaration=True)


def _inject_ignored_number_as_text(
  xlsx_path: Path,
  workbook: WorkbookType,
  sqref_by_sheet: dict[str, str],
) -> None:
  _postprocess_saved_xlsx(
    xlsx_path,
    workbook=workbook,
    ignored_text_sqref_by_sheet=sqref_by_sheet,
  )


def _excel_part_path(path: Path) -> Path:
  return path.with_name(f'.{path.stem}{path.suffix}.part')


def save_workbook_to_path(
  workbook: WorkbookType,
  path: Path,
  password: Optional[str] = None,
  *,
  ignored_text_sqref_by_sheet: dict[str, str] | None = None,
) -> None:
  path.parent.mkdir(parents=True, exist_ok=True)
  tmp_path = _excel_part_path(path)
  if tmp_path.exists():
    tmp_path.unlink(missing_ok=True)

  try:
    workbook.save(tmp_path)
    if ignored_text_sqref_by_sheet or password:
      _postprocess_saved_xlsx(
        tmp_path,
        workbook=workbook,
        ignored_text_sqref_by_sheet=ignored_text_sqref_by_sheet,
        password=password,
      )
    os.replace(tmp_path, path)
    tmp_path = None
  except OSError as exc:
    raise ExcelFileError(format_excel_error(exc)) from exc
  finally:
    if tmp_path is not None and tmp_path.exists():
      tmp_path.unlink(missing_ok=True)


def create_empty_workbook() -> WorkbookType:
  workbook = Workbook()
  default_sheet = workbook.active
  workbook.remove(default_sheet)
  return workbook
