from __future__ import annotations

import re
from pathlib import Path
from typing import List, Optional

from openpyxl.worksheet.worksheet import Worksheet

from app.services.excel_crypto import load_workbook_from_path
from app.services.excel_service import current_month_sheet_name


EXCEL_FILETYPES = [
  ('Excel', '*.xlsx;*.xlsm'),
  ('Todos', '*.*'),
]


def parse_headers_from_text(text: str) -> List[str]:
  """Interpreta linha(s) colada(s) do Excel ou texto separado por tab/;/|."""
  if not text or not str(text).strip():
    return []

  lines = [line.strip() for line in str(text).splitlines() if line.strip()]
  if not lines:
    return []

  candidate = max(lines, key=len)
  if '\t' in candidate:
    parts = candidate.split('\t')
  elif ';' in candidate:
    parts = candidate.split(';')
  elif '|' in candidate:
    parts = candidate.split('|')
  else:
    parts = re.split(r'\s{2,}', candidate)

  return [part.strip() for part in parts if part and str(part).strip()]


def _headers_from_sheet(sheet: Worksheet) -> List[str]:
  if sheet.max_row < 1:
    return []
  headers: List[str] = []
  for col in range(1, sheet.max_column + 1):
    value = sheet.cell(row=1, column=col).value
    if value is None:
      continue
    text = str(value).strip()
    if text:
      headers.append(text)
  return headers


def list_excel_sheets(path: Path, password: Optional[str] = None) -> List[str]:
  workbook = load_workbook_from_path(path, password=password)
  return list(workbook.sheetnames)


def read_headers_from_excel(
  excel_path: Path,
  *,
  password: Optional[str] = None,
  sheet_name: Optional[str] = None,
) -> List[str]:
  workbook = load_workbook_from_path(excel_path, password=password)
  target_sheet = (sheet_name or '').strip() or current_month_sheet_name()
  if target_sheet in workbook.sheetnames:
    sheet = workbook[target_sheet]
  elif workbook.sheetnames:
    sheet = workbook[workbook.sheetnames[0]]
  else:
    return []
  return _headers_from_sheet(sheet)
